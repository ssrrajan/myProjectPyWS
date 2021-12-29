# Copy range of cells as a nested list
# Takes: start cell, end cell, and sheet you want to copy from.
import openpyxl


def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    # Loops through selected Rows
    for i in range(startRow, endRow + 1, 1):
        # Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol, endCol + 1, 1):
            rowSelected.append(sheet.cell(row=i, column=j).value)
        # Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)

    return rangeSelected


def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
    countRow = 0
    for i in range(startRow, endRow + 1, 1):
        countCol = 0
        for j in range(startCol, endCol + 1, 1):
            sheetReceiving.cell(row=i, column=j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1


filename = "C:\\Sunder\\STQC.xlsx"
wb = openpyxl.load_workbook(filename)
sht1 = wb.get_sheet_by_name("STQC")
sht2name = "copytest"
print(wb.sheetnames)


# print(wb.sheetnames)
# wb.save(filename)


rowct = sht1.max_row
colct = sht1.max_column

print(rowct)
print(colct)

startCol = startRow = endCol = endRow = 1


for rc in range (2, rowct):
    if not sht1.cell(row = rc, column = 2).value == None:
        startRow = rc
        startCol = 1
        endCol = colct
        endRow = rc
        for enrc in range(rc, rowct+1):
            endRow = enrc
            if not sht1.cell(row=enrc + 1, column=1).value == None:
                break

        print("{}: {}".format(sht1.cell(row=rc, column=1).value, endRow))

        if sht2name in wb.sheetnames:
            wb.remove_sheet(wb.get_sheet_by_name(sht2name))
        sht2 = wb.create_sheet(sht2name)
        cprng = copyRange(1, 1, colct, 1, sht1)
        pasteRange(1, 1, colct, 1, sht2, cprng)
        cprng = copyRange(startCol, startRow, endCol, endRow, sht1)
        pasteRange(1, 2, endCol, endRow-rc+2, sht2, cprng)
        wb.save(filename)