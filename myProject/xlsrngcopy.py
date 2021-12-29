import openpyxl


def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []

    for i in range(startRow, endRow + 1, 1):
        rowSelected = []
        for j in range(startCol, endCol + 1, 1):
            rowSelected.append(sheet.cell(row=i, column=j).value)
        rangeSelected.append(rowSelected)

    return rangeSelected


def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
    countRow = 0
    for i in range(startRow, endRow + 1, 1):
        countCol = 0
        for j in range(startCol, endCol + 1, 1):
            sheetReceiving.cell(row=i, column=j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1


filename = "C:\\Sunder\\test.xlsx"
wb = openpyxl.load_workbook(filename)
sht = wb.get_sheet_by_name("Work packages")

rowct = sht.max_row
colct = sht.max_column

cprng = copyRange(1, 1, colct, 1, sht)

sht2 = wb.create_sheet("copytest")

pasteRange(1, 1, colct, 1, sht2, cprng)

cprng = copyRange(1, 5, colct, 5, sht)

pasteRange(1, 2, colct, 2, sht2, cprng)
wb.save(filename)


