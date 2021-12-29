import openpyxl

filename = "C:\\Sunder\\L1TestCases-WindowsUpd.xlsx"
wb = openpyxl.load_workbook(filename)
chksht = wb["Image_Quality_Testing"]
updsht = wb["L1 Test Cases"]


chkrowct = chksht.max_row
chkcolct = chksht.max_column

updrowct = updsht.max_row
updcolct = updsht.max_column



ct = 0
id = ""
# for i in range(2, updrowct+1):
#     if not updsht.cell(row=i,column=1).value == None:
#         phase = "NA"
#         for k in range(2, chkrowct+1):
#             if chksht.cell(row=k,column=1).value in updsht.cell(row=i,column=2).value:
#                 phase = chksht.cell(row=k,column=2).value
#                 break
#         updsht.cell(row=i, column=3).value = phase

for i in range(2, updrowct+1):
    if updsht.cell(row=i,column=1).value == None:
        id = ""
        for k in range(2, chkrowct+1):
            if not chksht.cell(row=k,column=1).value == None:
                if updsht.cell(row=i,column=2).value in chksht.cell(row=k,column=2).value:
                    id = chksht.cell(row=k,column=1).value
                    break
        updsht.cell(row=i, column=1).value = id
        ct = ct+1


wb.save(filename)
print(ct)