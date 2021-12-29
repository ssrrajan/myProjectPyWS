import time

from selenium import webdriver
from selenium.webdriver import ActionChains, Keys

from selenium.webdriver.common.by import By



driver = webdriver.Chrome(executable_path="C:\\chromedriver.exe")
driver.maximize_window()
driver.implicitly_wait(10)

import openpyxl


def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    # Loops through selected Rows
    for i in range(startRow, endRow + 1, 1):
        # Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol, endCol + 1, 1):
            rowSelected.append(sheet.cell(row=i, column=j).value)
        # Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)

    return rangeSelected


def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
    countRow = 0
    for i in range(startRow, endRow + 1, 1):
        countCol = 0
        for j in range(startCol, endCol + 1, 1):
            sheetReceiving.cell(row=i, column=j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1


filename = "C:\\Sunder\\STQC.xlsx"
wb = openpyxl.load_workbook(filename)

sht1 = wb.get_sheet_by_name("copytest")

rowct = sht1.max_row
colct = sht1.max_column



driver.get("http://rndprojects.precisionit.co.in/openproject/")
print("Loaded the URL for ")
username = "PG0046"
pwd = "WEL(0^^E"

driver.find_element(By.XPATH, value="//div[@id='login-form']/form/div[1]/div/span/input").send_keys(username)
driver.find_element(By.XPATH, value="//div[@id='login-form']/form/div[2]/div/span/input").send_keys(pwd)
driver.find_element(By.XPATH, value="//div[@id='login-form']/form/input[4]").click()

driver.find_element(By.LINK_TEXT, value="Select a project").click()
driver.find_element(By.LINK_TEXT, value="» InnaITAadhaar Test Repositories").click()
driver.find_element(By.ID, value="main-menu-work-packages").click()
driver.find_element(By.CLASS_NAME, value="wp-create-button").click()
wptypes = driver.find_elements(By.XPATH, value="//div[@id = 'types-context-menu']/ul/li")

print(len(wptypes))

for wp in wptypes:
    wpname = wp.text
    print(wpname)
    if wpname == "TEST CASE":
        wp.click()
        break
titletxt = sht1.cell(row=2, column=1).value + "_" + sht1.cell(row=2, column=2).value
desctxt = sht1.cell(row=2, column=3).value
driver.find_element(By.ID, value="wp-new-inline-edit--field-subject").send_keys(titletxt)
time.sleep(2)
driver.find_element(By.XPATH, value="//div[@class = 'inline-edit--active-field inplace-edit description']/form/ng-component/div/div/op-ckeditor/div/div[2]/div/p").clear()
time.sleep(2)
driver.find_element(By.XPATH, value="//div[@class = 'inline-edit--active-field inplace-edit description']/form/ng-component/div/div/op-ckeditor/div/div[2]/div/p").send_keys(desctxt)
cprng = copyRange(4, 1, colct, rowct, sht1)

print(len(cprng))

for i in range(1, len(cprng)+1):
    print(i)
    dat = cprng[i-1]
    for k in range(1, len(dat)+1):
        if not dat[k-1] == None:
            print(dat[k-1])
            driver.find_element(By.XPATH, value="{}{}{}{}{}".format("//div[@class = 'inline-edit--active-field inplace-edit customField13']/form/ng-component/div/div/op-ckeditor/div/div[2]/div/figure/table/tbody/tr[", i, "]/td[", k, "]")).send_keys(dat[k-1],Keys.TAB)
        else:
            driver.find_element(By.XPATH, value="{}{}{}{}{}".format("//div[@class = 'inline-edit--active-field inplace-edit customField13']/form/ng-component/div/div/op-ckeditor/div/div[2]/div/figure/table/tbody/tr[", i, "]/td[", k, "]")).send_keys(Keys.TAB)

driver.find_element(By.ID, value="work-packages--edit-actions-save").click()
wpid = driver.find_element(By.CLASS_NAME, value="work-packages--info-row").text
print(wpid.split(":")[0].strip("#"))
