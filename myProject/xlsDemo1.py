import openpyxl

book = openpyxl.load_workbook("C:\\Sunder\\pyxlDemo1.xlsx")
sheet = book.get_sheet_by_name("TestData")

cell = sheet.cell(row=1,column = 2)
print(cell.value)

# sheet.cell(row=2,column=5).value = "Male"
#
# print(sheet.cell(row=2,column=5).value)


print(sheet.max_row)
print(sheet.max_column)

for i in range(2,sheet.max_row+1):
    for k in range(2,sheet.max_column+1):
        print(sheet.cell(row=i,column=k).value)


for i in range(2,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value == "Test2":
        print(sheet.cell(row=i,column=1).value)
        for k in range(2,sheet.max_column+1):
            print(sheet.cell(row=i,column=k).value)