import openpyxl

book = openpyxl.load_workbook("C:\\Sunder\\pyxlDemo1.xlsx")
sheet = book.get_sheet_by_name("TestData")

mxrow = sheet.max_row
mxcol = sheet.max_column

Dict = {}
myList = []

for i in range(2, mxrow+1):
    Dict = {}
    for k in range(2, mxcol+1):
        Dict[sheet.cell(row=1, column=k).value] = sheet.cell(row=i, column=k).value

    print(Dict)
    myList.append(Dict)



print(myList)