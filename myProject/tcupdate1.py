import time

from selenium import webdriver
from selenium.webdriver import ActionChains, Keys

from selenium.webdriver.common.by import By



driver = webdriver.Chrome(executable_path="C:\\chromedriver.exe")
driver.maximize_window()
driver.implicitly_wait(10)

import openpyxl


def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    # Loops through selected Rows
    for i in range(startRow, endRow + 1, 1):
        # Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol, endCol + 1, 1):
            rowSelected.append(sheet.cell(row=i, column=j).value)
        # Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)

    return rangeSelected


def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
    countRow = 0
    for i in range(startRow, endRow + 1, 1):
        countCol = 0
        for j in range(startCol, endCol + 1, 1):
            sheetReceiving.cell(row=i, column=j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1


filename = "C:\\Sunder\\STQC.xlsx"
wb = openpyxl.load_workbook(filename)

sht1 = wb.get_sheet_by_name("copytest")

rowct = sht1.max_row
colct = sht1.max_column



driver.get("http://rndprojects.precisionit.co.in/openproject/projects/innaitaadhaar-test-repositories/work_packages/1586/activity")

username = "PG0046"
pwd = "WEL(0^^E"

driver.find_element(By.XPATH, value="//div[@id='login-form']/form/div[1]/div/span/input").send_keys(username)
driver.find_element(By.XPATH, value="//div[@id='login-form']/form/div[2]/div/span/input").send_keys(pwd)
driver.find_element(By.XPATH, value="//div[@id='login-form']/form/input[4]").click()


driver.get("http://rndprojects.precisionit.co.in/openproject/projects/innaitaadhaar-test-repositories/work_packages/1586/activity")
driver.find_element(By.XPATH, value="//div[@class = 'wp-breadcrumb -show']/ul/li/wp-breadcrumb-parent/accessible-by-keyboard/a").click()
driver.find_element(By.XPATH, value="//div[@class='wp-relations--autocomplete']/ng-select/div/div/div[2]/input").send_keys("1510", Keys.ENTER)
print(driver.find_element(By.XPATH, value="//div[@class = 'ng-dropdown-panel-items scroll-host']/div[2]/div").text)