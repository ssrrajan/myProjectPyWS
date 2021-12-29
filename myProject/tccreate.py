import time

from selenium import webdriver
from selenium.webdriver import ActionChains, Keys
from pymsgbox import *
from selenium.webdriver.common.by import By
import openpyxl

def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    # Loops through selected Rows
    for i in range(startRow, endRow + 1, 1):
        # Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol, endCol + 1, 1):
            rowSelected.append(sheet.cell(row=i, column=j).value)
        # Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)

    return rangeSelected

def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
    countRow = 0
    for i in range(startRow, endRow + 1, 1):
        countCol = 0
        for j in range(startCol, endCol + 1, 1):
            sheetReceiving.cell(row=i, column=j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

driverpath = "C:\\chromedriver.exe"
url = "http://rndprojects.precisionit.co.in/openproject/"
TestRepoTxt = "» PIVOT Test Repositories"
username = "PG0046"
pwd = "WEL(0^^E"
filename = "C:\\Sunder\\Smart Light BLE Test Case.xlsx"
shtname = "Template"
sht2name = "copytest"
wb = openpyxl.load_workbook(filename)
sht1 = wb.get_sheet_by_name(shtname)

rowct = sht1.max_row
colct = sht1.max_column

startCol = startRow = endCol = endRow = 1




print("Looking into Sheet provided")

ct = 0
avl = 0
navl = 0
for rc in range (2, rowct+1):
    if not sht1.cell(row=rc, column=2).value == None:
        ct = ct + 1
    if not sht1.cell(row=rc, column=1).value == None:
        avl = avl + 1
navl = ct-avl

print("Total {} Test Cases found...".format(ct))
print("{} Test Cases already available and {} to be created".format(avl, navl))


if navl > 0:

    print("Creating {} Test Cases in OpenProject".format(navl))
    driver = webdriver.Chrome(executable_path=driverpath)
    driver.maximize_window()
    driver.implicitly_wait(10)
    driver.get(url)
    print("Loaded the URL for Open Project......")
    driver.find_element(By.XPATH, value="//div[@id='login-form']/form/div[1]/div/span/input").send_keys(username)
    driver.find_element(By.XPATH, value="//div[@id='login-form']/form/div[2]/div/span/input").send_keys(pwd)
    driver.find_element(By.XPATH, value="//div[@id='login-form']/form/input[4]").click()
    print("Logged in as {} User......".format(username))
    driver.find_element(By.LINK_TEXT, value="Select a project").click()
    driver.find_element(By.LINK_TEXT, value=TestRepoTxt).click()
    driver.find_element(By.ID, value="main-menu-work-packages").click()
    print("Entered into InnaITAadhaar Test Repositories WorkPackages......")

    ct = 0
    for rc in range (2, rowct+1):

        if (not sht1.cell(row=rc, column=2).value == None) and (sht1.cell(row=rc, column=1).value == None):

            driver.find_element(By.CLASS_NAME, value="wp-create-button").click()
            wptypes = driver.find_elements(By.XPATH, value="//div[@id = 'types-context-menu']/ul/li")

            for wp in wptypes:
                wpname = wp.text
                if wpname == "TEST CASE":
                    wp.click()
                    break

            ct = ct + 1
            startRow = rc
            startCol = 1
            endCol = colct
            endRow = rc
            for enrc in range(rc, rowct+1):
                endRow = enrc
                if not sht1.cell(row=enrc + 1, column=2).value == None:
                    break

            if sht2name in wb.sheetnames:
                wb.remove_sheet(wb.get_sheet_by_name(sht2name))
            sht2 = wb.create_sheet(sht2name)
            cprng = copyRange(1, 1, colct, 1, sht1)
            pasteRange(1, 1, colct, 1, sht2, cprng)
            cprng = copyRange(startCol, startRow, endCol, endRow, sht1)
            pasteRange(1, 2, endCol, endRow-rc+2, sht2, cprng)
            wb.save(filename)

            titletxt = sht2.cell(row=2, column=2).value
            desctxt = sht2.cell(row=2, column=3).value
            pretxt = sht2.cell(row=2, column=4).value
            cprowct = sht2.max_row
            cpcolct = sht2.max_column

            cprng = copyRange(5, 1, cpcolct, cprowct, sht2)

            driver.find_element(By.ID, value = "wp-new-inline-edit--field-subject").send_keys(titletxt)
            time.sleep(2)
            driver.find_element(By.XPATH, value = "//div[@class = 'inline-edit--active-field inplace-edit description']/form/ng-component/div/div/op-ckeditor/div/div[2]/div/p").clear()
            time.sleep(1)
            driver.find_element(By.XPATH, value = "//div[@class = 'inline-edit--active-field inplace-edit description']/form/ng-component/div/div/op-ckeditor/div/div[2]/div/p").send_keys(desctxt)
            if not pretxt == None:
                # time.sleep(1)
                driver.find_element(By.XPATH, value = "//div[@class = 'inline-edit--active-field inplace-edit customField19']/form/ng-component/div/div/op-ckeditor/div/div[2]/div/p").send_keys(pretxt)

            for i in range(1, len(cprng) + 1):
                time.sleep(1)
                dat = cprng[i - 1]
                for k in range(1, len(dat)):
                    if not dat[k - 1] == None:
                        driver.find_element(By.XPATH, value="{}{}{}{}{}".format("//div[@class = 'inline-edit--active-field inplace-edit customField13']/form/ng-component/div/div/op-ckeditor/div/div[2]/div/figure/table/tbody/tr[", i, "]/td[", k, "]")).send_keys(dat[k - 1], Keys.TAB)
                    else:
                        driver.find_element(By.XPATH, value="{}{}{}{}{}".format("//div[@class = 'inline-edit--active-field inplace-edit customField13']/form/ng-component/div/div/op-ckeditor/div/div[2]/div/figure/table/tbody/tr[",i, "]/td[", k, "]")).send_keys(Keys.TAB)

            driver.find_element(By.ID, value="work-packages--edit-actions-save").click()
            wpid = driver.find_element(By.CLASS_NAME, value="work-packages--info-row").text.split(":")[0].strip("#")

            sht1.cell(row=rc, column=1).value = wpid

            print("{}: {}: {}".format(ct, sht1.cell(row=rc, column=2).value, sht1.cell(row=rc, column=1).value))
            wb.remove_sheet(wb.get_sheet_by_name(sht2name))
            wb.save(filename)
            if ct == 50:
                break
    print("Total {} Test Cases Created......".format(ct))
    driver.quit()
    alert(text="Total {} Test Cases Created......".format(ct), title='Test Case', button='OK')
else:
    alert(text="Total {} Test Cases already available......".format(avl), title='Test Case', button='OK')
